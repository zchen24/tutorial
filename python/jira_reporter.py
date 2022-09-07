#!/usr/bin/env python3

"""
Fetch Jira tickets and generate an Excel report

Author: Zihan Chen
Date: 2021-09-08

# Tested on Windows
# ==== Anaconda Shell ===
# pip install jira
# pip install openpyxl

# How to generate API token
# https://id.atlassian.com/manage-profile/security/api-tokens#
"""

import sys
from PyQt5.QtWidgets import *
from jira import JIRA
from openpyxl import Workbook, styles
import os

SERVER = "https://domain.atlassian.net/"
USER = "support@example.com"
TOKEN = '111111111111111111'


class JiraReporter():
    def __init__(self):
        self.jira = None

    def query_yellow_items(self, project):
        issues = self.jira.search_issues(
            'due >= "0" AND due <= 1w AND project = {} AND issuetype in (Bug, Candidate, Improvement, "New Feature", Story) AND status in (Backlog, "In Progress", "In Review", Interviewing, "On Hold", "Selected for Development", open) ORDER BY created DESC'.format(project),
            maxResults=100)
        return issues

    def query_green_items(self):
        pass

    def query_red_items(self):
        pass

    def export_excel(self, proj):
        if self.jira is None:
            options = {"server": SERVER}
            user = USER
            token = TOKEN
            self.jira = JIRA(options, basic_auth=(user, token))

        # generate excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'RNC'

        headers = ('#JIRA', 'PROJ', 'Task', 'DDL', 'Status')
        hdrs_width = (10, 15, 50, 15, 10)

        # COLOR_RED = "00FF0000"
        # COLOR_YEL = "00FFFF00"
        COLOR_GRN = "0000FF00"
        COLOR_HDR = "00DCDCDC"

        # header
        ws.row_dimensions[1].font = styles.Font(bold=True)
        for i, c in enumerate(headers):
            ws.cell(row=1, column=i + 1).value = headers[i]
            ws.cell(row=1, column=i + 1).fill = styles.PatternFill("solid", fgColor=COLOR_HDR)
            ws.column_dimensions[chr(ord('A') + i)].width = hdrs_width[i]

        # query
        issues = self.query_yellow_items(proj)
        
        for r, issue in enumerate(issues):
            ws.cell(row=2 + r, column=1).value = issue.key
            ws.cell(row=2 + r, column=1).hyperlink = "https://domain.atlassian.net/browse/{}".format(issue.key)
            ws.cell(row=2 + r, column=1).style = "Hyperlink"
            ws.cell(row=2 + r, column=2).value = issue.fields.project.key
            ws.cell(row=2 + r, column=3).value = issue.fields.summary
            ws.cell(row=2 + r, column=4).value = issue.fields.duedate  # TODO: format this
            ws.cell(row=2 + r, column=5).fill = styles.PatternFill("solid", fgColor=COLOR_GRN)

        # green

        # yellow

        # red

        wb.save('test.xlsx')
        print('Export complete')


class MyWidget(QWidget):
    def __init__(self):
        super(MyWidget, self).__init__()
        self.setWindowTitle('JIRA Reporter')
        self.setMinimumWidth(400)

        vbox = QVBoxLayout()
        grid = QGridLayout()
        # lbUser = QLabel("User: ")
        # self.leUser = QLineEdit()
        # lbToken = QLabel("Token: ")
        # self.leToken = QLineEdit()
        lbProject = QLabel("Project: ")
        self.leProject = QLineEdit()
        # grid.addWidget(lbUser, 0, 0)
        # grid.addWidget(self.leUser, 0, 1)
        # grid.addWidget(lbToken, 1, 0)
        # grid.addWidget(self.leToken, 1, 1)
        grid.addWidget(lbProject, 2, 0)
        grid.addWidget(self.leProject, 2, 1)

        # set defaults
        self.leProject.setText('RNC')

        pb1 = QPushButton('Generate')
        pb1.setCheckable(True)
        pb1.setChecked(True)   # select rb1

        pb2 = QPushButton('Open Excel')
        pb2.setCheckable(True)
        pb2.setStyleSheet("QPushButton:checked { background-color: green; }")

        vbox.addLayout(grid)
        vbox.addWidget(pb1)
        vbox.addWidget(pb2)
        self.setLayout(vbox)

        pb1.clicked.connect(self.slot_pb_generate)
        pb2.clicked.connect(self.slot_pb_open)

    def slot_pb_generate(self):
        print('Generating Excel Report')
        reporter = JiraReporter()
        reporter.export_excel(self.leProject.text())
        os.system('start "excel" "test.xlsx')

    def slot_pb_open(self):
        print('Opening Excel Report')
        os.system('start "excel" "test.xlsx')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mw = MyWidget()
    mw.show()
    sys.exit(app.exec_())
